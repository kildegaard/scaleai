import hashlib
import os
from datetime import datetime
import openpyxl


def calculate_checksum(file_path):
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        # Read and update the hash in chunks
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


if __name__ == "__main__":
    # Get the list of files
    files = os.listdir("./my_files")

    # Create a new Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the column headers
    worksheet["A1"], worksheet["B1"], worksheet["C1"] = (
        "File Name",
        "SHA-256 Checksum",
        "Generated At",
    )

    # Generate and write checksums for the files
    row = 2
    for file in files:
        file_path = os.path.join(os.getcwd(), "my_files", file)
        checksum = calculate_checksum(file_path)
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        worksheet.cell(row=row, column=1, value=file)
        worksheet.cell(row=row, column=2, value=checksum)
        worksheet.cell(row=row, column=3, value=generated_at)

        row += 1

    # Save the workbook to an Excel file
    workbook.save("file_checksums.xlsx")
    print("File checksums saved to file_checksums.xlsx")
